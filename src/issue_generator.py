"""问题清单生成模块 - 输出Excel格式"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from typing import List, Dict
from pathlib import Path
import time

class IssueListGenerator:
    """问题清单生成器"""

    def __init__(self, output_dir: str):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def generate_excel(self, issues: List[Dict], filename: str = None) -> str:
        """生成Excel问题清单"""
        if not filename:
            filename = f"问题清单_{time.strftime('%Y%m%d_%H%M%S')}.xlsx"

        filepath = self.output_dir / filename

        # 创建DataFrame
        df = pd.DataFrame(issues)
        df.insert(0, '序号', range(1, len(df) + 1))

        # 重命名列
        column_map = {
            'company': '二级公司',
            'project': '项目名称',
            'city': '城市',
            'field': '数据异常项',
            'description': '问题描述',
            'suggestion': '修改意见',
            'priority': '问题优先级'
        }
        df = df.rename(columns=column_map)

        # 选择需要的列
        cols = ['序号', '二级公司', '项目名称', '城市', '数据异常项', '问题描述', '修改意见', '问题优先级']
        df = df[[c for c in cols if c in df.columns]]

        # 写入Excel
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='问题清单', index=False)
            self._format_excel(writer.book['问题清单'], df)

        return str(filepath)

    def _format_excel(self, sheet, df):
        """格式化Excel样式"""
        # 表头样式
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')

        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 高优先级标红
        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        priority_col = None
        for idx, col in enumerate(df.columns, 1):
            if col == '问题优先级':
                priority_col = idx
                break

        if priority_col:
            for row in range(2, len(df) + 2):
                if sheet.cell(row, priority_col).value == '高':
                    for col in range(1, len(df.columns) + 1):
                        sheet.cell(row, col).fill = red_fill

        # 调整列宽
        for col in sheet.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            sheet.column_dimensions[col_letter].width = min(max_length + 2, 50)

